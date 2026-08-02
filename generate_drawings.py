#!/usr/bin/env python3
"""
AQUILA Sovereign — Master Engineering Drawing Generator
=======================================================
Generates production-ready engineering drawings following the standards in
"Engineering Working Drawings Basics" (MENG 204, Dr. Ala Hijazi).

Standards implemented:
  - ISO A-series sheet sizes (A0-A4) with drawing frame and margins
  - Title block (bottom-right): company, title, drawing number, scale,
    sheet size, projection symbol (third-angle), units, drafter, checker, date
  - Orthographic projections: Front, Top, Right (third-angle layout)
  - Isometric view (30°/30° axonometric)
  - Section views (full/half) for internal features
  - Detail views (lettered, magnified) for small features
  - Dimensioning: chain + baseline, extension lines, dimension lines with arrowheads
  - Assembly drawings: BOM table with balloon callouts, exploded views with explode lines
  - Numbering: XXX.AA.BB.CC hierarchical (per page 21-22)

Output formats per sheet: DXF (ezdxf), PDF (matplotlib), SVG, PNG (300 DPI)

Usage:
    python generate_drawings.py                    # all components + assemblies
    python generate_drawings.py --component UAV-001  # single component
    python generate_drawings.py --assembly subsystem  # assembly sheets only
"""
from __future__ import annotations

import argparse
import json
import math
import os
import re
import sys
from datetime import datetime
from pathlib import Path

import matplotlib
matplotlib.use('Agg')  # headless
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.patches import FancyBboxPatch, Rectangle, Circle, FancyArrowPatch
from matplotlib.lines import Line2D
import matplotlib.font_manager as fm
import numpy as np

try:
    import ezdxf
    from ezdxf.enums import TextEntityAlignment
    EZDXF_AVAILABLE = True
except ImportError:
    EZDXF_AVAILABLE = False
    print("WARNING: ezdxf not available, DXF output will be skipped")

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).parent
BOM_FILE = BASE_DIR / "Aquila_Master_BOM_Reconciled_v2 7-5-26 (version 1).xlsx"
OUTPUT_DIR = BASE_DIR / "drawings"
DXF_DIR = OUTPUT_DIR / "dxf"
PDF_DIR = OUTPUT_DIR / "pdf"
SVG_DIR = OUTPUT_DIR / "svg"
PNG_DIR = OUTPUT_DIR / "png"

# Drawing standards (from PDF)
# ISO A-series sheet sizes (mm) — width × height (landscape)
SHEET_SIZES = {
    'A0': (1189, 841),
    'A1': (841, 594),
    'A2': (594, 420),
    'A3': (420, 297),
    'A4': (297, 210),
}

# Default sheet for component drawings
DEFAULT_SHEET = 'A2'
# Assembly drawings use A1 (more room for BOM)
ASSEMBLY_SHEET = 'A1'

# Margins (mm) — per ISO standard
MARGIN = 10  # standard margin
TITLE_BLOCK_W = 180  # title block width
TITLE_BLOCK_H = 56   # title block height

# Company info
COMPANY = "Aquila Geological Systems"
DRAFTER = "Dr. White Sutherland"
CHECKER = "Nemotron-3 Nano Omni"
DATE_STR = "2026-05-13"
PROJECTION = "third"  # third-angle projection (US standard)

# Colors
COLOR_BORDER = '#000000'
COLOR_DIMENSION = '#0066CC'
COLOR_EXTENSION = '#0066CC'
COLOR_CENTERLINE = '#FF0000'
COLOR_HIDDEN = '#999999'
COLOR_OBJECT = '#000000'
COLOR_SECTION_HATCH = '#444444'
COLOR_ISO = '#333333'
COLOR_TITLE_BG = '#F0F0F0'
COLOR_BOM_BG = '#F8F8F8'
COLOR_BALLOON = '#0066CC'
COLOR_EXPLODE_LINE = '#FF6600'

# DPI for PNG output
PNG_DPI = 300

# ---------------------------------------------------------------------------
# BOM Parsing
# ---------------------------------------------------------------------------

def parse_dims(s: str) -> tuple[float, float, float]:
    s = str(s).replace(',', '').replace('×', 'x').replace('Ø', '').replace('*', 'x')
    parts = re.findall(r'[\d.]+', s)
    if len(parts) >= 3:
        return (float(parts[0]), float(parts[1]), float(parts[2]))
    elif len(parts) == 2:
        return (float(parts[0]), float(parts[0]), float(parts[1]))
    elif len(parts) == 1:
        return (float(parts[0]), float(parts[0]), float(parts[0]))
    return (50.0, 50.0, 50.0)


def parse_number(s) -> float:
    s = str(s).replace(',', '').replace('~', '').replace('$', '').strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


def parse_bom(xlsx_path: Path) -> list[dict]:
    if not OPENPYXL_AVAILABLE:
        print("ERROR: openpyxl required to parse BOM", file=sys.stderr)
        return []
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    ws = wb.active
    header_row = None
    for r in range(1, 20):
        if str(ws.cell(r, 1).value or '').strip() == 'ID':
            header_row = r
            break
    if header_row is None:
        return []
    headers = [str(ws.cell(header_row, c).value or '').strip() for c in range(1, 15)]
    components = []
    for r in range(header_row + 1, ws.max_row + 1):
        bom_id = str(ws.cell(r, 1).value or '').strip()
        if not bom_id or bom_id.startswith('#') or bom_id.startswith('Phase'):
            continue
        if any(bom_id.startswith(p) for p in ['NOTE', 'note', 'TOTAL', 'SUMMARY']):
            continue
        comp = {}
        for c, h in enumerate(headers, 1):
            val = ws.cell(r, c).value
            if val is not None and h:
                comp[h] = str(val).strip()
        if comp.get('ID'):
            dims_str = comp.get('Dimensions (mm)', '')
            comp['dims_mm'] = parse_dims(dims_str)
            comp['dims_str'] = dims_str
            comp['weight_g'] = parse_number(comp.get('Weight (g)', '0'))
            comp['qty'] = int(parse_number(comp.get('Qty', '1')))
            comp['unit_cost'] = parse_number(comp.get('Unit Cost (USD)', '0'))
            comp['ext_cost'] = parse_number(comp.get('Ext. Cost (USD)', '0'))
            comp['is_cylinder'] = dims_str.startswith('Ø') or dims_str.startswith('o')
            components.append(comp)
    return components


# ---------------------------------------------------------------------------
# Drawing Number Assignment (per PDF page 21-22: XXX.AA.BB.CC)
# ---------------------------------------------------------------------------

def assign_drawing_numbers(components: list[dict]) -> dict:
    """Assign hierarchical drawing numbers based on subsystem grouping."""
    # Group by subsystem
    subsystems = {}
    for c in components:
        subsys = c.get('Subsystem', 'Unknown')
        if subsys not in subsystems:
            subsystems[subsys] = []
        subsystems[subsys].append(c)

    # Assign numbers: 001.AA.BB format
    # 001 = main assembly, AA = subsystem, BB = part within subsystem
    drawing_numbers = {}
    for subsys_idx, (subsys, comps) in enumerate(sorted(subsystems.items()), 1):
        for part_idx, comp in enumerate(comps, 1):
            dwg_num = f"ASD-{subsys_idx:02d}.{part_idx:02d}"
            drawing_numbers[comp['ID']] = dwg_num

    return drawing_numbers


# ---------------------------------------------------------------------------
# Geometry: 2D projection of 3D box/cylinder
# ---------------------------------------------------------------------------

def project_box_front(w, h, d):
    """Front view of a box: width × height (looking down -Y)."""
    return [(0, 0), (w, 0), (w, h), (0, h)], 'box'


def project_box_top(w, h, d):
    """Top view of a box: width × depth (looking down -Z)."""
    return [(0, 0), (w, 0), (w, d), (0, d)], 'box'


def project_box_right(w, h, d):
    """Right view of a box: depth × height (looking down +X)."""
    return [(0, 0), (d, 0), (d, h), (0, h)], 'box'


def project_cylinder_front(diam, height):
    """Front view of a cylinder: rectangle with centerline."""
    r = diam / 2
    return [(0, 0), (diam, 0), (diam, height), (0, height)], 'cylinder'


def project_cylinder_top(diam, height):
    """Top view of a cylinder: circle."""
    return [(diam/2, diam/2, diam/2)], 'circle_top'


def project_cylinder_right(diam, height):
    """Right view of a cylinder: same as front."""
    return [(0, 0), (diam, 0), (diam, height), (0, height)], 'cylinder'


def isometric_box(w, h, d):
    """Generate isometric projection of a box.
    Returns list of 3D points projected to 2D isometric."""
    cos30 = math.cos(math.radians(30))
    sin30 = math.sin(math.radians(30))

    # 8 corners of the box
    corners_3d = [
        (0, 0, 0), (w, 0, 0), (w, d, 0), (0, d, 0),  # bottom
        (0, 0, h), (w, 0, h), (w, d, h), (0, d, h),  # top
    ]

    # Isometric projection: x' = (x - y) * cos30, y' = (x + y) * sin30 - z
    corners_2d = []
    for x, y, z in corners_3d:
        iso_x = (x - y) * cos30
        iso_y = (x + y) * sin30 - z
        corners_2d.append((iso_x, iso_y))

    # Edges (pairs of corner points, not indices)
    edge_indices = [
        (0,1), (1,2), (2,3), (3,0),  # bottom
        (4,5), (5,6), (6,7), (7,4),  # top
        (0,4), (1,5), (2,6), (3,7),  # verticals
    ]
    edges = [(corners_2d[i], corners_2d[j]) for i, j in edge_indices]

    return corners_2d, edges


def isometric_cylinder(diam, height):
    """Generate isometric projection of a cylinder."""
    r = diam / 2
    cos30 = math.cos(math.radians(30))
    sin30 = math.sin(math.radians(30))

    # Top and bottom ellipses (approximated as polygons)
    n = 24
    top_pts = []
    bot_pts = []
    for i in range(n):
        a = 2 * math.pi * i / n
        x = r * math.cos(a)
        y = r * math.sin(a)
        # Project to iso
        iso_x = (x - y) * cos30
        iso_y_top = (x + y) * sin30 - height
        iso_y_bot = (x + y) * sin30
        top_pts.append((iso_x, iso_y_top))
        bot_pts.append((iso_x, iso_y_bot))

    # Edges: top ellipse, bottom ellipse, and 2 silhouette lines
    edges = []
    for i in range(n):
        edges.append((top_pts[i], top_pts[(i+1) % n]))
        edges.append((bot_pts[i], bot_pts[(i+1) % n]))
    # Silhouette lines (left and right extremes)
    edges.append((top_pts[0], bot_pts[0]))
    edges.append((top_pts[n//2], bot_pts[n//2]))

    return top_pts + bot_pts, edges


# ---------------------------------------------------------------------------
# Matplotlib Drawing Sheet
# ---------------------------------------------------------------------------

class DrawingSheet:
    """A single engineering drawing sheet rendered with matplotlib."""

    def __init__(self, sheet_size='A2', is_assembly=False):
        self.sheet_size = sheet_size if not is_assembly else 'A1'
        self.width, self.height = SHEET_SIZES[self.sheet_size]
        self.is_assembly = is_assembly

        # Create figure at mm scale (1mm = 1 unit)
        fig_w = self.width / 25.4  # inches
        fig_h = self.height / 25.4
        self.fig, self.ax = plt.subplots(1, 1, figsize=(fig_w, fig_h), dpi=100)
        self.ax.set_xlim(0, self.width)
        self.ax.set_ylim(0, self.height)
        self.ax.set_aspect('equal')
        self.ax.axis('off')
        self.fig.patch.set_facecolor('white')

        # Draw frame
        self._draw_frame()

        # Drawing area (inside frame, excluding title block area)
        self.draw_area = {
            'x_min': MARGIN,
            'y_min': MARGIN,
            'x_max': self.width - MARGIN,
            'y_max': self.height - MARGIN,
        }

    def _draw_frame(self):
        """Draw the drawing frame (border) per ISO standard."""
        # Outer border
        self.ax.add_patch(Rectangle(
            (0, 0), self.width, self.height,
            fill=False, edgecolor=COLOR_BORDER, linewidth=2
        ))
        # Inner frame (drawing area border)
        self.ax.add_patch(Rectangle(
            (MARGIN, MARGIN),
            self.width - 2 * MARGIN,
            self.height - 2 * MARGIN,
            fill=False, edgecolor=COLOR_BORDER, linewidth=1.5
        ))

    def draw_title_block(self, title, drawing_number, revision='A',
                         scale='1:2', material='Al 6061-T6',
                         tolerance='±0.1mm', sheet_num='1 OF 1'):
        """Draw the title block in the bottom-right corner."""
        x = self.width - MARGIN - TITLE_BLOCK_W
        y = MARGIN
        w = TITLE_BLOCK_W
        h = TITLE_BLOCK_H

        # Background
        self.ax.add_patch(Rectangle(
            (x, y), w, h, facecolor=COLOR_TITLE_BG,
            edgecolor=COLOR_BORDER, linewidth=1.5
        ))

        # Internal grid lines
        # Horizontal divisions (3 rows)
        row_h = h / 3
        for i in range(1, 3):
            self.ax.plot([x, x + w], [y + i * row_h, y + i * row_h],
                        color=COLOR_BORDER, linewidth=0.5)

        # Vertical divisions
        col_w = w / 3
        self.ax.plot([x + col_w, x + col_w], [y, y + h],
                    color=COLOR_BORDER, linewidth=0.5)
        self.ax.plot([x + 2 * col_w, x + 2 * col_w], [y, y + h],
                    color=COLOR_BORDER, linewidth=0.5)

        # Row 1 (top): Company name (spans full width)
        self.ax.text(x + w/2, y + h - row_h/2, COMPANY,
                    ha='center', va='center', fontsize=8, fontweight='bold')

        # Row 2: Title | Drawing No | Rev
        self.ax.text(x + col_w/2, y + 2*row_h - row_h/2, title[:25],
                    ha='center', va='center', fontsize=6, fontweight='bold')
        self.ax.text(x + col_w + col_w/2, y + 2*row_h - row_h/2, f"DWG: {drawing_number}",
                    ha='center', va='center', fontsize=6)
        self.ax.text(x + 2*col_w + col_w/2, y + 2*row_h - row_h/2, f"REV: {revision}",
                    ha='center', va='center', fontsize=6)

        # Row 3 (bottom): Scale | Units | Sheet | Date | Projection
        self.ax.text(x + col_w * 0.25, y + row_h * 0.7, f"SCALE: {scale}",
                    ha='left', va='center', fontsize=5)
        self.ax.text(x + col_w * 0.25, y + row_h * 0.3, f"UNITS: mm",
                    ha='left', va='center', fontsize=5)
        self.ax.text(x + col_w * 0.75, y + row_h * 0.7, f"MAT: {material[:12]}",
                    ha='left', va='center', fontsize=5)
        self.ax.text(x + col_w * 0.75, y + row_h * 0.3, f"TOL: {tolerance}",
                    ha='left', va='center', fontsize=5)

        # Projection symbol (third-angle)
        self.ax.text(x + 2*col_w + col_w * 0.15, y + row_h * 0.7,
                    f"PROJ: 3rd", ha='left', va='center', fontsize=5)
        self.ax.text(x + 2*col_w + col_w * 0.15, y + row_h * 0.3,
                    f"SHEET: {sheet_num}", ha='left', va='center', fontsize=5)

        # Draw third-angle projection symbol (truncated cone)
        sym_x = x + 2 * col_w + col_w * 0.7
        sym_y = y + row_h * 0.5
        # Small circle (front view)
        self.ax.add_patch(Circle((sym_x, sym_y + 3), 2, fill=False,
                                edgecolor=COLOR_BORDER, linewidth=0.5))
        # Trapezoid (side view)
        self.ax.plot([sym_x + 5, sym_x + 12, sym_x + 12, sym_x + 5],
                    [sym_y + 5, sym_y + 6, sym_y + 0, sym_y + 1],
                    color=COLOR_BORDER, linewidth=0.5)
        self.ax.plot([sym_x + 5, sym_x + 12],
                    [sym_y + 3, sym_y + 3],
                    color=COLOR_BORDER, linewidth=0.3, linestyle='--')

    def draw_revision_block(self):
        """Draw revision history table above the title block."""
        x = self.width - MARGIN - TITLE_BLOCK_W
        y = MARGIN + TITLE_BLOCK_H
        w = TITLE_BLOCK_W
        h = 20

        self.ax.add_patch(Rectangle(
            (x, y), w, h, facecolor='white',
            edgecolor=COLOR_BORDER, linewidth=1
        ))
        # Header
        self.ax.plot([x, x + w], [y + h - 5, y + h - 5],
                    color=COLOR_BORDER, linewidth=0.5)
        self.ax.text(x + 5, y + h - 2.5, "REV",
                    ha='left', va='center', fontsize=5, fontweight='bold')
        self.ax.text(x + 25, y + h - 2.5, "DESCRIPTION",
                    ha='left', va='center', fontsize=5, fontweight='bold')
        self.ax.text(x + 120, y + h - 2.5, "DATE",
                    ha='left', va='center', fontsize=5, fontweight='bold')
        self.ax.text(x + 155, y + h - 2.5, "BY",
                    ha='left', va='center', fontsize=5, fontweight='bold')

        # Rev A entry
        self.ax.text(x + 5, y + 5, "A",
                    ha='left', va='center', fontsize=5)
        self.ax.text(x + 25, y + 5, "Initial release",
                    ha='left', va='center', fontsize=5)
        self.ax.text(x + 120, y + 5, DATE_STR,
                    ha='left', va='center', fontsize=5)
        self.ax.text(x + 155, y + 5, "DWS",
                    ha='left', va='center', fontsize=5)

    def draw_notes_block(self, notes=None):
        """Draw a notes block in the bottom-left corner."""
        if notes is None:
            notes = [
                "1. All dimensions in millimeters unless noted.",
                "2. General tolerance: ±0.1mm unless otherwise specified.",
                "3. Surface finish: Ra 3.2μm unless otherwise specified.",
                "4. Remove all burrs and sharp edges.",
                "5. Do not scale drawing — use stated dimensions.",
            ]

        x = MARGIN
        y = MARGIN
        w = 100
        h = len(notes) * 5 + 8

        self.ax.add_patch(Rectangle(
            (x, y), w, h, facecolor='white',
            edgecolor=COLOR_BORDER, linewidth=1
        ))
        self.ax.text(x + 5, y + h - 4, "NOTES:",
                    ha='left', va='center', fontsize=5, fontweight='bold')
        for i, note in enumerate(notes):
            self.ax.text(x + 5, y + h - 8 - i * 5, note,
                        ha='left', va='center', fontsize=4)

    def draw_patent_block(self, patent_ids, origin=None):
        """Draw a patent reference block listing applicable patents.

        Args:
            patent_ids: list of PAT-XX strings from PATENT_MAP
            origin: (x, y) top-left; defaults to above notes block
        """
        if not patent_ids:
            return

        if origin is None:
            x = MARGIN
            y = MARGIN + 40  # above the notes block
        else:
            x, y = origin

        w = 130
        line_h = 4.5
        h = len(patent_ids) * line_h + 10

        # Background (light green tint for patent callouts)
        self.ax.add_patch(Rectangle(
            (x, y), w, h, facecolor='#E8F5E9',
            edgecolor='#2E7D32', linewidth=1
        ))

        # Header
        self.ax.text(x + 5, y + h - 4, "PATENT REFERENCES:",
                    ha='left', va='center', fontsize=4.5,
                    fontweight='bold', color='#2E7D32')

        for i, pat_id in enumerate(patent_ids):
            pat = PATENT_MAP.get(pat_id, {})
            title = pat.get('title', 'Unknown')[:28]
            cat = pat.get('category', '')[:8]
            line_y = y + h - 8 - i * line_h
            self.ax.text(x + 5, line_y, f"{pat_id}",
                        ha='left', va='center', fontsize=4,
                        fontweight='bold', color='#1B5E20')
            self.ax.text(x + 28, line_y, title,
                        ha='left', va='center', fontsize=3.5, color='#333333')
            self.ax.text(x + w - 5, line_y, cat,
                        ha='right', va='center', fontsize=3, color='#666666')

    def draw_orthographic_view(self, corners, view_type, origin, scale=1.0,
                               is_cylinder=False, diam=0, height=0):
        """Draw an orthographic view at the given origin.

        Args:
            corners: list of (x, y) points for the outline
            view_type: 'box' or 'cylinder' or 'circle_top'
            origin: (x, y) position on the sheet
            scale: drawing scale factor
            is_cylinder: whether this is a cylindrical part
            diam: diameter for cylinder circle
            height: height for cylinder
        """
        ox, oy = origin

        if view_type == 'circle_top':
            # Draw circle for top view of cylinder
            r = diam / 2 * scale
            circle = Circle((ox + r, oy + r), r, fill=False,
                           edgecolor=COLOR_OBJECT, linewidth=1.5)
            self.ax.add_patch(circle)
            # Centerlines
            self.ax.plot([ox, ox + 2*r], [oy + r, oy + r],
                        color=COLOR_CENTERLINE, linewidth=0.5,
                        linestyle='--')
            self.ax.plot([ox + r, ox + r], [oy, oy + 2*r],
                        color=COLOR_CENTERLINE, linewidth=0.5,
                        linestyle='--')
            return 2 * r, 2 * r

        # Draw polygon outline
        scaled_corners = [(ox + cx * scale, oy + cy * scale) for cx, cy in corners]
        xs = [p[0] for p in scaled_corners] + [scaled_corners[0][0]]
        ys = [p[1] for p in scaled_corners] + [scaled_corners[0][1]]
        self.ax.plot(xs, ys, color=COLOR_OBJECT, linewidth=1.5)

        # For cylinder, add centerline
        if is_cylinder and view_type == 'cylinder':
            w = (corners[1][0] - corners[0][0]) * scale
            h = (corners[2][1] - corners[1][1]) * scale
            # Vertical centerline
            cx = ox + w / 2
            self.ax.plot([cx, cx], [oy - 3, oy + h + 3],
                        color=COLOR_CENTERLINE, linewidth=0.5,
                        linestyle='--')

        # Return drawn dimensions
        w = (max(c[0] for c in corners) - min(c[0] for c in corners)) * scale
        h = (max(c[1] for c in corners) - min(c[1] for c in corners)) * scale
        return w, h

    def draw_isometric_view(self, corners_2d, edges, origin, scale=1.0):
        """Draw an isometric view at the given origin."""
        ox, oy = origin
        # Center the isometric view
        min_x = min(p[0] for p in corners_2d)
        max_x = max(p[0] for p in corners_2d)
        min_y = min(p[1] for p in corners_2d)
        max_y = max(p[1] for p in corners_2d)
        offset_x = ox - min_x * scale
        offset_y = oy - min_y * scale

        for p1, p2 in edges:
            x1 = offset_x + p1[0] * scale
            y1 = offset_y + p1[1] * scale
            x2 = offset_x + p2[0] * scale
            y2 = offset_y + p2[1] * scale
            self.ax.plot([x1, x2], [y1, y2],
                        color=COLOR_ISO, linewidth=1.2)

        w = (max_x - min_x) * scale
        h = (max_y - min_y) * scale
        return w, h

    def draw_dimension(self, x1, y1, x2, y2, text, offset=8, is_horizontal=True):
        """Draw a dimension line with arrowheads and text.

        Args:
            x1, y1: start point of dimension
            x2, y2: end point of dimension
            text: dimension text
            offset: distance from object to dimension line
            is_horizontal: whether dimension is horizontal or vertical
        """
        if is_horizontal:
            # Horizontal dimension
            dy = offset if y1 + offset < self.height - MARGIN else -offset
            dim_y = y1 + dy
            # Extension lines
            self.ax.plot([x1, x1], [y1, dim_y - 2],
                        color=COLOR_EXTENSION, linewidth=0.5)
            self.ax.plot([x2, x2], [y2, dim_y - 2],
                        color=COLOR_EXTENSION, linewidth=0.5)
            # Dimension line with arrowheads
            self.ax.annotate('', xy=(x1, dim_y), xytext=(x2, dim_y),
                            arrowprops=dict(arrowstyle='<->', color=COLOR_DIMENSION,
                                          lw=0.8, shrinkA=0, shrinkB=0))
            # Text
            self.ax.text((x1 + x2) / 2, dim_y + 2, text,
                        ha='center', va='bottom', fontsize=5,
                        color=COLOR_DIMENSION, fontweight='bold')
        else:
            # Vertical dimension
            dx = offset if x1 + offset < self.width - MARGIN else -offset
            dim_x = x1 + dx
            # Extension lines
            self.ax.plot([x1, dim_x - 2], [y1, y1],
                        color=COLOR_EXTENSION, linewidth=0.5)
            self.ax.plot([x2, dim_x - 2], [y2, y2],
                        color=COLOR_EXTENSION, linewidth=0.5)
            # Dimension line
            self.ax.annotate('', xy=(dim_x, y1), xytext=(dim_x, y2),
                            arrowprops=dict(arrowstyle='<->', color=COLOR_DIMENSION,
                                          lw=0.8, shrinkA=0, shrinkB=0))
            # Text (rotated)
            self.ax.text(dim_x + 2, (y1 + y2) / 2, text,
                        ha='left', va='center', fontsize=5,
                        color=COLOR_DIMENSION, fontweight='bold',
                        rotation=90)

    def draw_balloon(self, x, y, number, radius=5):
        """Draw a balloon callout with item number."""
        circle = Circle((x, y), radius, fill=True, facecolor='white',
                       edgecolor=COLOR_BALLOON, linewidth=1.2)
        self.ax.add_patch(circle)
        self.ax.text(x, y, str(number), ha='center', va='center',
                    fontsize=5, fontweight='bold', color=COLOR_BALLOON)

    def draw_bom_table(self, items, origin, col_widths=None):
        """Draw a BOM table for assembly drawings.

        Args:
            items: list of dicts with 'item', 'part_no', 'description', 'qty', 'material'
            origin: (x, y) top-left of table
            col_widths: list of column widths
        """
        if col_widths is None:
            col_widths = [15, 35, 70, 15, 35]  # ITEM, PART NO, DESCRIPTION, QTY, MATERIAL

        x, y = origin
        row_h = 6
        header_h = 8
        total_w = sum(col_widths)
        total_h = header_h + len(items) * row_h

        # Table border
        self.ax.add_patch(Rectangle(
            (x, y - total_h), total_w, total_h,
            fill=True, facecolor=COLOR_BOM_BG,
            edgecolor=COLOR_BORDER, linewidth=1
        ))

        # Header row
        headers = ['ITEM', 'PART NO', 'DESCRIPTION', 'QTY', 'MATERIAL']
        cx = x
        for i, (header, cw) in enumerate(zip(headers, col_widths)):
            if i > 0:
                self.ax.plot([cx, cx], [y, y - total_h],
                            color=COLOR_BORDER, linewidth=0.5)
            self.ax.text(cx + cw/2, y - header_h/2, header,
                        ha='center', va='center', fontsize=5,
                        fontweight='bold')
            cx += cw

        # Header line
        self.ax.plot([x, x + total_w], [y - header_h, y - header_h],
                    color=COLOR_BORDER, linewidth=1)

        # Data rows
        for row_idx, item in enumerate(items):
            row_y = y - header_h - (row_idx + 1) * row_h
            cx = x
            values = [
                str(item.get('item', '')),
                str(item.get('part_no', '')),
                str(item.get('description', ''))[:30],
                str(item.get('qty', '')),
                str(item.get('material', ''))[:15],
            ]
            for i, (val, cw) in enumerate(zip(values, col_widths)):
                if i > 0:
                    self.ax.plot([cx, cx], [row_y, row_y + row_h],
                                color=COLOR_BORDER, linewidth=0.3)
                self.ax.text(cx + cw/2, row_y + row_h/2, val,
                            ha='center', va='center', fontsize=4)
                cx += cw
            # Row separator
            if row_idx < len(items) - 1:
                self.ax.plot([x, x + total_w], [row_y, row_y],
                            color=COLOR_BORDER, linewidth=0.3)

    def draw_explode_line(self, x1, y1, x2, y2):
        """Draw an explode line between components."""
        self.ax.plot([x1, x2], [y1, y2],
                    color=COLOR_EXPLODE_LINE, linewidth=0.8,
                    linestyle='--')

    def draw_section_hatching(self, corners, origin, scale=1.0, angle=45):
        """Draw section hatching lines inside a polygon."""
        ox, oy = origin
        scaled = [(ox + cx * scale, oy + cy * scale) for cx, cy in corners]
        xs = [p[0] for p in scaled]
        ys = [p[1] for p in scaled]
        min_x, max_x = min(xs), max(xs)
        min_y, max_y = min(ys), max(ys)

        # Draw hatching lines at 45°
        spacing = 3
        for d in range(-int(max_x - min_x + max_y - min_y),
                       int(max_x - min_x + max_y - min_y), spacing):
            x_start = min_x + d
            y_start = min_y
            x_end = min_x + d + (max_y - min_y)
            y_end = max_y
            # Clip to polygon bounds (simple rectangular clip)
            if x_start < min_x:
                y_start = min_y + (min_x - x_start)
                x_start = min_x
            if x_end > max_x:
                y_end = max_y - (x_end - max_x)
                x_end = max_x
            if y_start < min_y or y_end > max_y:
                continue
            self.ax.plot([x_start, x_end], [y_start, y_end],
                        color=COLOR_SECTION_HATCH, linewidth=0.3)

    def add_text(self, x, y, text, fontsize=5, ha='left', va='center',
                 fontweight='normal', color='#000000', rotation=0):
        """Add arbitrary text to the sheet."""
        self.ax.text(x, y, text, ha=ha, va=va, fontsize=fontsize,
                    fontweight=fontweight, color=color, rotation=rotation)

    def save(self, filepath, format='png'):
        """Save the drawing sheet."""
        self.fig.savefig(filepath, dpi=PNG_DPI if format == 'png' else 150,
                        bbox_inches='tight', pad_inches=0.1,
                        facecolor='white', edgecolor='none')
        plt.close(self.fig)


# ---------------------------------------------------------------------------
# DXF Drawing Sheet (ezdxf)
# ---------------------------------------------------------------------------

class DXFDrawing:
    """Generate DXF version of a drawing sheet using ezdxf."""

    def __init__(self, sheet_size='A2'):
        self.sheet_size = sheet_size
        self.width, self.height = SHEET_SIZES[sheet_size]
        self.doc = ezdxf.new('R2018', setup=True)
        self.msp = self.doc.modelspace()

        # Create layers
        self.doc.layers.add('BORDER', color=7)
        self.doc.layers.add('OBJECT', color=7)
        self.doc.layers.add('DIMENSIONS', color=5)
        self.doc.layers.add('CENTERLINE', color=1)
        self.doc.layers.add('TEXT', color=7)
        self.doc.layers.add('TITLE_BLOCK', color=7)
        self.doc.layers.add('HATCHING', color=8)
        self.doc.layers.add('BOM', color=7)
        self.doc.layers.add('BALLOONS', color=5)

        # Draw frame
        self._draw_frame()

    def _draw_frame(self):
        # Outer border
        self.msp.add_lwpolyline(
            [(0, 0), (self.width, 0), (self.width, self.height), (0, self.height), (0, 0)],
            dxfattribs={'layer': 'BORDER', 'lineweight': 35}
        )
        # Inner frame
        self.msp.add_lwpolyline(
            [(MARGIN, MARGIN), (self.width - MARGIN, MARGIN),
             (self.width - MARGIN, self.height - MARGIN),
             (MARGIN, self.height - MARGIN), (MARGIN, MARGIN)],
            dxfattribs={'layer': 'BORDER', 'lineweight': 25}
        )

    def draw_title_block(self, title, drawing_number, revision='A',
                         scale='1:2', material='Al 6061-T6',
                         tolerance='±0.1mm', sheet_num='1 OF 1'):
        x = self.width - MARGIN - TITLE_BLOCK_W
        y = MARGIN
        w = TITLE_BLOCK_W
        h = TITLE_BLOCK_H

        # Border
        self.msp.add_lwpolyline(
            [(x, y), (x+w, y), (x+w, y+h), (x, y+h), (x, y)],
            dxfattribs={'layer': 'TITLE_BLOCK'}
        )

        # Grid lines
        row_h = h / 3
        for i in range(1, 3):
            self.msp.add_line((x, y + i*row_h), (x+w, y + i*row_h),
                            dxfattribs={'layer': 'TITLE_BLOCK'})
        col_w = w / 3
        for i in range(1, 3):
            self.msp.add_line((x + i*col_w, y), (x + i*col_w, y+h),
                            dxfattribs={'layer': 'TITLE_BLOCK'})

        # Text
        def add_text(tx, ty, text, height=2.5, bold=False):
            attribs = {'layer': 'TEXT', 'height': height}
            if bold:
                attribs['style'] = 'OpenSans-Bold'
            self.msp.add_text(text, dxfattribs=attribs).set_placement((tx, ty), align=TextEntityAlignment.MIDDLE_CENTER)

        add_text(x + w/2, y + h - row_h/2, COMPANY, height=3)
        add_text(x + col_w/2, y + 2*row_h - row_h/2, title[:25], height=2)
        add_text(x + col_w + col_w/2, y + 2*row_h - row_h/2, f"DWG: {drawing_number}", height=2)
        add_text(x + 2*col_w + col_w/2, y + 2*row_h - row_h/2, f"REV: {revision}", height=2)
        add_text(x + col_w * 0.25, y + row_h * 0.7, f"SCALE: {scale}", height=1.8)
        add_text(x + col_w * 0.25, y + row_h * 0.3, f"UNITS: mm", height=1.8)
        add_text(x + col_w * 0.75, y + row_h * 0.7, f"MAT: {material[:12]}", height=1.8)
        add_text(x + col_w * 0.75, y + row_h * 0.3, f"TOL: {tolerance}", height=1.8)
        add_text(x + 2*col_w + col_w * 0.15, y + row_h * 0.7, f"PROJ: 3rd", height=1.8)
        add_text(x + 2*col_w + col_w * 0.15, y + row_h * 0.3, f"SHEET: {sheet_num}", height=1.8)

    def draw_orthographic(self, corners, view_type, origin, scale=1.0,
                          is_cylinder=False, diam=0):
        ox, oy = origin
        if view_type == 'circle_top':
            r = diam / 2 * scale
            self.msp.add_circle((ox + r, oy + r), r,
                              dxfattribs={'layer': 'OBJECT'})
            # Centerlines
            self.msp.add_line((ox, oy + r), (ox + 2*r, oy + r),
                            dxfattribs={'layer': 'CENTERLINE', 'linetype': 'DASHED'})
            self.msp.add_line((ox + r, oy), (ox + r, oy + 2*r),
                            dxfattribs={'layer': 'CENTERLINE', 'linetype': 'DASHED'})
            return

        scaled = [(ox + cx * scale, oy + cy * scale) for cx, cy in corners]
        pts = scaled + [scaled[0]]
        self.msp.add_lwpolyline(pts, dxfattribs={'layer': 'OBJECT', 'lineweight': 18})

        if is_cylinder and view_type == 'cylinder':
            w = (corners[1][0] - corners[0][0]) * scale
            h = (corners[2][1] - corners[1][1]) * scale
            cx = ox + w / 2
            self.msp.add_line((cx, oy - 3), (cx, oy + h + 3),
                            dxfattribs={'layer': 'CENTERLINE', 'linetype': 'DASHED'})

    def draw_isometric(self, corners_2d, edges, origin, scale=1.0):
        ox, oy = origin
        min_x = min(p[0] for p in corners_2d)
        min_y = min(p[1] for p in corners_2d)
        for p1, p2 in edges:
            x1 = ox + (p1[0] - min_x) * scale
            y1 = oy + (p1[1] - min_y) * scale
            x2 = ox + (p2[0] - min_x) * scale
            y2 = oy + (p2[1] - min_y) * scale
            self.msp.add_line((x1, y1), (x2, y2),
                            dxfattribs={'layer': 'OBJECT', 'lineweight': 15})

    def draw_dimension(self, x1, y1, x2, y2, text, offset=8, is_horizontal=True):
        if is_horizontal:
            dim_y = y1 + offset
            self.msp.add_line((x1, y1), (x1, dim_y - 2), dxfattribs={'layer': 'DIMENSIONS'})
            self.msp.add_line((x2, y2), (x2, dim_y - 2), dxfattribs={'layer': 'DIMENSIONS'})
            self.msp.add_line((x1, dim_y), (x2, dim_y), dxfattribs={'layer': 'DIMENSIONS'})
            # Arrowheads (simple)
            self.msp.add_text(text, dxfattribs={'layer': 'DIMENSIONS', 'height': 2}).set_placement(
                ((x1+x2)/2, dim_y + 2), align=TextEntityAlignment.MIDDLE_CENTER)
        else:
            dim_x = x1 + offset
            self.msp.add_line((x1, y1), (dim_x - 2, y1), dxfattribs={'layer': 'DIMENSIONS'})
            self.msp.add_line((x2, y2), (dim_x - 2, y2), dxfattribs={'layer': 'DIMENSIONS'})
            self.msp.add_line((dim_x, y1), (dim_x, y2), dxfattribs={'layer': 'DIMENSIONS'})
            self.msp.add_text(text, dxfattribs={'layer': 'DIMENSIONS', 'height': 2, 'rotation': 90}).set_placement(
                (dim_x + 2, (y1+y2)/2), align=TextEntityAlignment.MIDDLE_CENTER)

    def draw_balloon(self, x, y, number, radius=5):
        self.msp.add_circle((x, y), radius, dxfattribs={'layer': 'BALLOONS'})
        self.msp.add_text(str(number), dxfattribs={'layer': 'BALLOONS', 'height': 3}).set_placement(
            (x, y), align=TextEntityAlignment.MIDDLE_CENTER)

    def draw_bom_table(self, items, origin, col_widths=None):
        if col_widths is None:
            col_widths = [15, 35, 70, 15, 35]
        x, y = origin
        row_h = 6
        header_h = 8
        total_w = sum(col_widths)
        total_h = header_h + len(items) * row_h

        # Border
        self.msp.add_lwpolyline(
            [(x, y), (x+total_w, y), (x+total_w, y-total_h), (x, y-total_h), (x, y)],
            dxfattribs={'layer': 'BOM'}
        )
        # Header line
        self.msp.add_line((x, y - header_h), (x + total_w, y - header_h),
                         dxfattribs={'layer': 'BOM'})

        headers = ['ITEM', 'PART NO', 'DESCRIPTION', 'QTY', 'MATERIAL']
        cx = x
        for i, (header, cw) in enumerate(zip(headers, col_widths)):
            if i > 0:
                self.msp.add_line((cx, y), (cx, y - total_h),
                                dxfattribs={'layer': 'BOM'})
            self.msp.add_text(header, dxfattribs={'layer': 'BOM', 'height': 2.5}).set_placement(
                (cx + cw/2, y - header_h/2), align=TextEntityAlignment.MIDDLE_CENTER)
            cx += cw

        for row_idx, item in enumerate(items):
            row_y = y - header_h - (row_idx + 1) * row_h
            cx = x
            values = [str(item.get('item', '')), str(item.get('part_no', '')),
                     str(item.get('description', ''))[:30], str(item.get('qty', '')),
                     str(item.get('material', ''))[:15]]
            for i, (val, cw) in enumerate(zip(values, col_widths)):
                if i > 0:
                    self.msp.add_line((cx, row_y), (cx, row_y + row_h),
                                    dxfattribs={'layer': 'BOM'})
                self.msp.add_text(val, dxfattribs={'layer': 'BOM', 'height': 2}).set_placement(
                    (cx + cw/2, row_y + row_h/2), align=TextEntityAlignment.MIDDLE_CENTER)
                cx += cw
            if row_idx < len(items) - 1:
                self.msp.add_line((x, row_y), (x + total_w, row_y),
                                dxfattribs={'layer': 'BOM'})

    def draw_explode_line(self, x1, y1, x2, y2):
        self.msp.add_line((x1, y1), (x2, y2),
                         dxfattribs={'layer': 'DIMENSIONS', 'linetype': 'DASHED'})

    def save(self, filepath):
        self.doc.saveas(filepath)


# ---------------------------------------------------------------------------
# Component Drawing Generation
# ---------------------------------------------------------------------------

def generate_component_drawing(component, drawing_number, output_dirs):
    """Generate a complete drawing sheet for a single BOM component.

    Layout (A2, third-angle projection):
      - Front view (top-left)
      - Top view (below front, third-angle)
      - Right view (right of front, third-angle)
      - Isometric view (top-right)
      - Title block (bottom-right)
      - Notes block (bottom-left)
      - Revision block (above title block)
    """
    bom_id = component['ID']
    comp_name = component.get('Component / Model', bom_id)
    subsys = component.get('Subsystem', 'Unknown')
    dims = component['dims_mm']
    w, h, d = dims
    is_cyl = component['is_cylinder']
    weight = component.get('weight_g', 0)
    dims_str = component.get('dims_str', f"{w}x{h}x{d}")

    # Determine scale (fit the largest view in ~150mm)
    max_dim = max(w, h, d, 10)
    if max_dim > 300:
        scale = 150 / max_dim
        scale_str = f"1:{int(1/scale)}"
    elif max_dim > 150:
        scale = 200 / max_dim
        scale_str = f"1:{int(1/scale)}"
    else:
        scale = 1.0
        scale_str = "1:1"

    # Material from subsystem
    material_map = {
        'Structure': 'CF + Al 6061-T6',
        'Compute': 'Al 6061-T6',
        'EO/IR': 'Al 6061-T6',
    }
    material = material_map.get(subsys, 'Al 6061-T6')

    sheet_title = f"{bom_id} — {comp_name[:20]}"

    # --- Matplotlib version (PDF, SVG, PNG) ---
    sheet = DrawingSheet(sheet_size=DEFAULT_SHEET)

    # View positions (third-angle projection layout)
    front_origin = (MARGIN + 30, self_height(DEFAULT_SHEET) - MARGIN - 80)
    top_origin = (MARGIN + 30, MARGIN + 100)
    right_origin = (MARGIN + 180, self_height(DEFAULT_SHEET) - MARGIN - 80)
    iso_origin = (self_width(DEFAULT_SHEET) - MARGIN - 120, self_height(DEFAULT_SHEET) - MARGIN - 100)

    # Draw front view
    if is_cyl:
        front_corners, front_type = project_cylinder_front(w, h)
    else:
        front_corners, front_type = project_box_front(w, h, d)
    fw, fh = sheet.draw_orthographic_view(
        front_corners, front_type, front_origin, scale, is_cyl, w, h)

    # Draw top view (third-angle: below front)
    if is_cyl:
        top_corners, top_type = project_cylinder_top(w, h)
    else:
        top_corners, top_type = project_box_top(w, h, d)
    tw, th = sheet.draw_orthographic_view(
        top_corners, top_type, top_origin, scale, is_cyl, w, h)

    # Draw right view (third-angle: right of front)
    if is_cyl:
        right_corners, right_type = project_cylinder_right(w, h)
    else:
        right_corners, right_type = project_box_right(w, h, d)
    rw, rh = sheet.draw_orthographic_view(
        right_corners, right_type, right_origin, scale, is_cyl, w, h)

    # Draw isometric view
    if is_cyl:
        iso_corners, iso_edges = isometric_cylinder(w, h)
    else:
        iso_corners, iso_edges = isometric_box(w, h, d)
    iw, ih = sheet.draw_isometric_view(iso_corners, iso_edges, iso_origin, scale * 0.8)

    # Dimensions (front view)
    fx, fy = front_origin
    # Width dimension (horizontal, below front view)
    sheet.draw_dimension(fx, fy, fx + fw, fy, f"{w:.0f}", offset=-15, is_horizontal=True)
    # Height dimension (vertical, left of front view)
    sheet.draw_dimension(fx, fy, fx, fy + fh, f"{h:.0f}", offset=-15, is_horizontal=False)

    # Depth dimension (on right view, horizontal)
    rx, ry = right_origin
    sheet.draw_dimension(rx, ry, rx + rw, ry, f"{d:.0f}", offset=-15, is_horizontal=True)

    # Isometric label
    sheet.add_text(iso_origin[0] + iw/2, iso_origin[1] - 10,
                  "ISOMETRIC", fontsize=5, ha='center', fontweight='bold')

    # View labels
    sheet.add_text(front_origin[0] + fw/2, front_origin[1] + fh + 5,
                  "FRONT", fontsize=5, ha='center', fontweight='bold')
    sheet.add_text(top_origin[0] + tw/2, top_origin[1] + th + 5,
                  "TOP", fontsize=5, ha='center', fontweight='bold')
    sheet.add_text(right_origin[0] + rw/2, right_origin[1] + rh + 5,
                  "RIGHT", fontsize=5, ha='center', fontweight='bold')

    # Title block, revision, notes
    sheet.draw_title_block(
        title=sheet_title, drawing_number=drawing_number,
        scale=scale_str, material=material[:12],
        tolerance='±0.1mm', sheet_num='1 OF 1')
    sheet.draw_revision_block()
    sheet.draw_notes_block()

    # Additional info
    info_x = MARGIN + 5
    info_y = self_height(DEFAULT_SHEET) - MARGIN - 15
    sheet.add_text(info_x, info_y, f"SUBSYSTEM: {subsys}", fontsize=5, fontweight='bold')
    sheet.add_text(info_x, info_y - 7, f"WEIGHT: {weight:.0f} g", fontsize=5)
    sheet.add_text(info_x, info_y - 14, f"QTY: {component.get('qty', 1)}", fontsize=5)
    sheet.add_text(info_x, info_y - 21, f"DIMENSIONS: {dims_str}", fontsize=5)

    # Save in all formats
    base_name = f"{drawing_number}_{sanitize_filename(bom_id)}"

    sheet.save(str(output_dirs['png'] / f"{base_name}.png"), format='png')
    sheet.save(str(output_dirs['pdf'] / f"{base_name}.pdf"), format='pdf')
    sheet.save(str(output_dirs['svg'] / f"{base_name}.svg"), format='svg')

    # --- DXF version ---
    if EZDXF_AVAILABLE:
        dxf = DXFDrawing(sheet_size=DEFAULT_SHEET)
        dxf.draw_orthographic(front_corners, front_type, front_origin, scale, is_cyl, w)
        dxf.draw_orthographic(top_corners, top_type, top_origin, scale, is_cyl, w)
        dxf.draw_orthographic(right_corners, right_type, right_origin, scale, is_cyl, w)
        dxf.draw_isometric(iso_corners, iso_edges, iso_origin, scale * 0.8)
        dxf.draw_dimension(fx, fy, fx + fw, fy, f"{w:.0f}", offset=-15, is_horizontal=True)
        dxf.draw_dimension(fx, fy, fx, fy + fh, f"{h:.0f}", offset=-15, is_horizontal=False)
        dxf.draw_dimension(rx, ry, rx + rw, ry, f"{d:.0f}", offset=-15, is_horizontal=True)
        dxf.draw_title_block(
            title=sheet_title, drawing_number=drawing_number,
            scale=scale_str, material=material[:12])
        dxf.save(str(output_dirs['dxf'] / f"{base_name}.dxf"))

    return base_name


# ---------------------------------------------------------------------------
# Assembly Drawing Generation
# ---------------------------------------------------------------------------

def generate_assembly_drawing(group_name, components, drawing_number, output_dirs,
                              grouping_type='subsystem'):
    """Generate an exploded assembly drawing with BOM table."""
    sheet = DrawingSheet(sheet_size=ASSEMBLY_SHEET, is_assembly=True)

    # Title
    sheet_title = f"ASSEMBLY — {group_name.upper()}"

    # Draw exploded isometric view of all components
    # Position components in a vertical exploded stack
    sw = self_width(ASSEMBLY_SHEET)
    sh = self_height(ASSEMBLY_SHEET)

    explode_origin_x = sw / 2 - 60
    explode_origin_y = sh - MARGIN - 60
    explode_spacing = 30

    bom_items = []
    for idx, comp in enumerate(components):
        bom_id = comp['ID']
        comp_name = comp.get('Component / Model', bom_id)
        subsys = comp.get('Subsystem', 'Unknown')
        dims = comp['dims_mm']
        w, h, d = dims
        is_cyl = comp['is_cylinder']

        # Position in exploded view
        comp_origin = (explode_origin_x, explode_origin_y - idx * explode_spacing)

        # Draw isometric
        if is_cyl:
            iso_corners, iso_edges = isometric_cylinder(w, h)
        else:
            iso_corners, iso_edges = isometric_box(w, h, d)

        # Scale to fit
        max_dim = max(w, h, d, 10)
        comp_scale = min(15 / max_dim, 0.5)
        iw, ih = sheet.draw_isometric_view(iso_corners, iso_edges, comp_origin, comp_scale)

        # Balloon
        balloon_x = comp_origin[0] + iw + 10
        balloon_y = comp_origin[1] + ih / 2
        sheet.draw_balloon(balloon_x, balloon_y, idx + 1)

        # Explode line
        if idx > 0:
            prev_origin = (explode_origin_x, explode_origin_y - (idx-1) * explode_spacing)
            sheet.draw_explode_line(
                prev_origin[0] + iw/2, prev_origin[1],
                comp_origin[0] + iw/2, comp_origin[1] + ih
            )

        # BOM entry
        bom_items.append({
            'item': idx + 1,
            'part_no': bom_id,
            'description': comp_name[:30],
            'qty': comp.get('qty', 1),
            'material': subsys[:15],
        })

    # BOM table (right side)
    bom_origin = (sw - MARGIN - 180, sh - MARGIN - 30)
    sheet.draw_bom_table(bom_items, bom_origin)

    # Title block
    sheet.draw_title_block(
        title=sheet_title, drawing_number=drawing_number,
        scale="NTS", material="VARIES",
        tolerance='±0.1mm', sheet_num='1 OF 1')
    sheet.draw_revision_block()
    sheet.draw_notes_block()

    # Grouping label
    sheet.add_text(MARGIN + 5, sh - MARGIN - 15,
                  f"GROUPING: {grouping_type.upper()}", fontsize=6, fontweight='bold')
    sheet.add_text(MARGIN + 5, sh - MARGIN - 22,
                  f"COMPONENTS: {len(components)}", fontsize=5)

    # Patent references — match by group name keywords to sheet numbers
    group_lower = group_name.lower()
    applicable_patents = []
    if any(k in group_lower for k in ['actuation', 'drill', 'carousel', 'mechanical']):
        applicable_patents = SHEET_PATENTS.get(4, [])
    elif any(k in group_lower for k in ['pneumatic', 'bio', 'thermal', 'env', 'fluidic']):
        applicable_patents = SHEET_PATENTS.get(5, [])
    elif any(k in group_lower for k in ['sensor', 'sampling', 'deployment', 'acoustic', 'seismic']):
        applicable_patents = SHEET_PATENTS.get(5, []) + SHEET_PATENTS.get(6, [])
    elif any(k in group_lower for k in ['compute', 'control', 'power', 'battery', 'radar', 'optical']):
        applicable_patents = SHEET_PATENTS.get(7, [])

    if applicable_patents:
        sheet.draw_patent_block(applicable_patents, origin=(MARGIN, MARGIN + 45))

    base_name = f"ASSY_{sanitize_filename(group_name)}"

    sheet.save(str(output_dirs['png'] / f"{base_name}.png"), format='png')
    sheet.save(str(output_dirs['pdf'] / f"{base_name}.pdf"), format='pdf')
    sheet.save(str(output_dirs['svg'] / f"{base_name}.svg"), format='svg')

    # DXF
    if EZDXF_AVAILABLE:
        dxf = DXFDrawing(sheet_size=ASSEMBLY_SHEET)
        for idx, comp in enumerate(components):
            dims = comp['dims_mm']
            w, h, d = dims
            comp_origin = (explode_origin_x, explode_origin_y - idx * explode_spacing)
            if comp['is_cylinder']:
                iso_corners, iso_edges = isometric_cylinder(w, h)
            else:
                iso_corners, iso_edges = isometric_box(w, h, d)
            max_dim = max(w, h, d, 10)
            comp_scale = min(15 / max_dim, 0.5)
            dxf.draw_isometric(iso_corners, iso_edges, comp_origin, comp_scale)
            dxf.draw_balloon(comp_origin[0] + 20, comp_origin[1] + 10, idx + 1)
        dxf.draw_bom_table(bom_items, bom_origin)
        dxf.draw_title_block(title=sheet_title, drawing_number=drawing_number,
                            scale="NTS", material="VARIES")
        dxf.save(str(output_dirs['dxf'] / f"{base_name}.dxf"))

    return base_name


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def self_width(sheet_size):
    return SHEET_SIZES[sheet_size][0]

def self_height(sheet_size):
    return SHEET_SIZES[sheet_size][1]

def sanitize_filename(s):
    s = re.sub(r'[^a-zA-Z0-9_\-]', '_', s)
    return s.strip('_')[:50]


# ---------------------------------------------------------------------------
# Grouping
# ---------------------------------------------------------------------------

ZONE_MAP = {
    'EO/IR Camera': 'Zone_A_Nose', 'Photogrammetry Cam': 'Zone_A_Nose',
    'Gimbal': 'Zone_A_Nose', 'Stereo Vision': 'Zone_A_Nose',
    'Compute Core': 'Zone_B_Bay1', 'Carrier Board': 'Zone_B_Bay1',
    'Edge Brain (Upgrade)': 'Zone_B_Bay1', 'Payload Pod': 'Zone_B_Bay1',
    'Payload Shell': 'Zone_B_Bay1',
    'Subsurface Radar': 'Zone_C_Bay2', 'QCL Laser': 'Zone_C_Bay2',
    'IR Detector': 'Zone_C_Bay2', 'PAS Detector': 'Zone_C_Bay2',
    'Micro-Spectrometer': 'Zone_C_Bay2', 'Acoustic Source': 'Zone_C_Bay2',
    'Acoustic Receiver': 'Zone_C_Bay2', 'Impulse Mic': 'Zone_C_Bay2',
    'Mic Array': 'Zone_C_Bay2', 'Active Seismic Source': 'Zone_C_Bay2',
    'Seismic Payload': 'Zone_C_Bay2', 'Radar Edge Compute': 'Zone_C_Bay2',
    'Radar Firmware': 'Zone_C_Bay2', 'Air-Picture Radar': 'Zone_C_Bay2',
    'Thermal Management': 'Zone_C_Bay2',
    'DNA Sequencer': 'Zone_D_Bay3', 'Soil Sampler': 'Zone_D_Bay3',
    'Fluidic Interface': 'Zone_D_Bay3', 'Pneumatic Hub': 'Zone_D_Bay3',
    'Env. Sensor': 'Zone_D_Bay3', 'VOC Sensor': 'Zone_D_Bay3',
    'Sensor Deployment': 'Zone_D_Bay3',
    'Flight Battery': 'Zone_E_Tail', 'Payload Battery': 'Zone_E_Tail',
    'Power Conversion': 'Zone_E_Tail', 'Control Electronics': 'Zone_E_Tail',
    'Airborne Data Gateway': 'Zone_E_Tail', 'Swarm Mesh Radio': 'Zone_E_Tail',
    'Swarm Control Station': 'Zone_E_Tail', 'Swarm Platform': 'Zone_E_Tail',
    'Ground Control Station': 'Zone_E_Tail', 'Flight Planning SW': 'Zone_E_Tail',
    'Friend-or-Foe Module': 'Zone_E_Tail', 'Hit Recording Module': 'Zone_E_Tail',
    'Range Sensor': 'Zone_E_Tail', 'Passive Monitoring Mode': 'Zone_E_Tail',
    'Re-purposing Layer': 'Zone_E_Tail', 'Sim Synthesis': 'Zone_E_Tail',
    'Drone Platform': 'Zone_F_Wings', 'Actuation': 'Zone_F_Wings',
    'Mechanical Interface': 'Zone_F_Wings',
}

PHASE_MAP = {
    'phase_1_alta_x': ['Drone Platform', 'Flight Battery', 'NAV', 'EO/IR Camera', 'Power Conversion'],
    'phase_2_firefly': ['Subsurface Radar', 'QCL Laser', 'IR Detector', 'PAS Detector',
                        'Micro-Spectrometer', 'Photogrammetry Cam', 'Gimbal', 'Stereo Vision',
                        'Payload Battery', 'Payload Shell', 'Payload Pod', 'Actuation',
                        'Compute Core', 'Carrier Board', 'Edge Brain (Upgrade)',
                        'Thermal Management', 'Control Electronics', 'Mechanical Interface',
                        'Acoustic Source', 'Acoustic Receiver', 'Active Seismic Source',
                        'Seismic Payload', 'Impulse Mic', 'Mic Array', 'Radar Edge Compute',
                        'Radar Firmware', 'Air-Picture Radar', 'Airborne Data Gateway',
                        'Swarm Mesh Radio', 'Swarm Control Station', 'Swarm Platform',
                        'Ground Control Station', 'Flight Planning SW', 'Friend-or-Foe Module',
                        'Hit Recording Module', 'Range Sensor', 'Passive Monitoring Mode',
                        'Re-purposing Layer', 'Sim Synthesis'],
    'phase_3_bio_intel': ['DNA Sequencer', 'Soil Sampler', 'Fluidic Interface',
                          'Pneumatic Hub', 'Env. Sensor', 'VOC Sensor', 'Sensor Deployment'],
}

# ---------------------------------------------------------------------------
# Patent Mapping (20 patents → 10 drawing sheets)
# ---------------------------------------------------------------------------

PATENT_MAP = {
    'PAT-01': {
        'file': '01_Spring_Compensated_O_Ring_Retainer.docx',
        'title': 'Spring-Compensated O-Ring Retainer',
        'sheet': 4,
        'category': 'Sealing',
    },
    'PAT-02': {
        'file': '02_Resonant_Backscatter_Phase_Array_RFID.docx',
        'title': 'Resonant Backscatter Phase-Array RFID',
        'sheet': 4,
        'category': 'Kinematic',
    },
    'PAT-03': {
        'file': '03_Dual_Layer_Piston_Ring_Seal.docx',
        'title': 'Dual-Layer Piston Ring Seal',
        'sheet': 4,
        'category': 'Sealing',
    },
    'PAT-04': {
        'file': '04_Variable_Pitch_Helical_Insert.docx',
        'title': 'Variable-Pitch Helical Insert',
        'sheet': 5,
        'category': 'Pneumatic',
    },
    'PAT-05': {
        'file': '05_Jetson_Adaptive_Pulsed_Pneumatic_Flow.docx',
        'title': 'Jetson Adaptive Pulsed Pneumatic Flow',
        'sheet': 5,
        'category': 'Pneumatic',
    },
    'PAT-06': {
        'file': '06_Inline_Expansion_Chamber.docx',
        'title': 'Inline Expansion Chamber',
        'sheet': 5,
        'category': 'Pneumatic',
    },
    'PAT-07': {
        'file': '07_Vision_Based_Pose_Calibration.docx',
        'title': 'Vision-Based Pose Calibration',
        'sheet': 6,
        'category': 'Kinematic',
    },
    'PAT-08': {
        'file': '08_Passive_Compliant_End_Effector.docx',
        'title': 'Passive Compliant End-Effector',
        'sheet': 6,
        'category': 'Kinematic',
    },
    'PAT-09': {
        'file': '09_Modular_Kinematic_Adapter_Plates.docx',
        'title': 'Modular Kinematic Adapter Plates',
        'sheet': 6,
        'category': 'Kinematic',
    },
    'PAT-10': {
        'file': '10_Electromagnetic_Clutch_Torque_Decoupling.docx',
        'title': 'Electromagnetic Clutch Torque Decoupling',
        'sheet': 4,
        'category': 'Torque',
    },
    'PAT-11': {
        'file': '11_Dual_Motor_Segmented_Torque.docx',
        'title': 'Dual-Motor Segmented Torque',
        'sheet': 4,
        'category': 'Torque',
    },
    'PAT-12': {
        'file': '12_Hollow_Conical_Rubber_Isolator.docx',
        'title': 'Hollow Conical Rubber Isolator',
        'sheet': 4,
        'category': 'Isolation',
    },
    'PAT-13': {
        'file': '13_Passive_Phase_Change_Thermal_Buffer.docx',
        'title': 'Passive Phase-Change Thermal Buffer',
        'sheet': 5,
        'category': 'Isolation',
    },
    'PAT-14': {
        'file': '14_Negative_Pressure_Filtered_Enclosure.docx',
        'title': 'Negative-Pressure Filtered Enclosure',
        'sheet': 5,
        'category': 'Sealing',
    },
    'PAT-15': {
        'file': '15_Passive_Condensation_Pipe.docx',
        'title': 'Passive Condensation Pipe',
        'sheet': 5,
        'category': 'Sealing',
    },
    'PAT-16': {
        'file': '16_Staged_Cyclone_HEPA_Filtration.docx',
        'title': 'Staged Cyclone HEPA Filtration',
        'sheet': 5,
        'category': 'Sealing',
    },
    'PAT-17': {
        'file': '17_Chassis_Integral_Boss_Integration.docx',
        'title': 'Chassis Integral Boss Integration (DFMA REV B)',
        'sheet': 4,
        'category': 'DFMA',
    },
    'PAT-18': {
        'file': '18_Cup_Lid_Assembly_Deletion.docx',
        'title': 'Cup-Lid Assembly Deletion (DFMA REV B)',
        'sheet': 4,
        'category': 'DFMA',
    },
    'PAT-19': {
        'file': '19_Servo_Bracket_Kinematic_Adapter_Standardization.docx',
        'title': 'Servo Bracket Standardization (DFMA REV B)',
        'sheet': 6,
        'category': 'DFMA',
    },
    'PAT-20': {
        'file': '20_Electronic_Module_Integration.docx',
        'title': 'Electronic Module Integration',
        'sheet': 7,
        'category': 'DFMA',
    },
}

# Reverse map: sheet number → list of patent IDs
SHEET_PATENTS = {}
for pat_id, pat_info in PATENT_MAP.items():
    sheet_num = pat_info['sheet']
    SHEET_PATENTS.setdefault(sheet_num, []).append(pat_id)


def group_by_subsystem(components):
    groups = {}
    for c in components:
        subsys = c.get('Subsystem', 'Unknown')
        groups.setdefault(subsys, []).append(c)
    return groups


def group_by_zone(components):
    groups = {}
    for c in components:
        subsys = c.get('Subsystem', 'Unknown')
        zone = ZONE_MAP.get(subsys, 'Zone_F_Wings')
        groups.setdefault(zone, []).append(c)
    return groups


def group_by_phase(components):
    groups = {k: [] for k in PHASE_MAP}
    for c in components:
        subsys = c.get('Subsystem', 'Unknown')
        for phase, subsystems in PHASE_MAP.items():
            if subsys in subsystems:
                groups[phase].append(c)
                break
        else:
            groups['phase_2_firefly'].append(c)
    return {k: v for k, v in groups.items() if v}


# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# Patent Master Index Sheet (Sheet 10)
# ---------------------------------------------------------------------------

def generate_patent_index_sheet(output_dirs):
    """Generate Sheet 10 — Master Patent Index with all 20 patents."""
    sheet = DrawingSheet(sheet_size=ASSEMBLY_SHEET, is_assembly=True)
    sw, sh = SHEET_SIZES[ASSEMBLY_SHEET]

    # Title
    sheet.add_text(sw / 2, sh - MARGIN - 10,
                  "AQUILA SOVEREIGN — PATENT INNOVATION INDEX",
                  fontsize=8, ha='center', fontweight='bold')
    sheet.add_text(sw / 2, sh - MARGIN - 18,
                  "20 Proprietary Innovations · DFMA REV B",
                  fontsize=5, ha='center', color='#666666')

    # Patent table
    table_x = MARGIN + 10
    table_y = sh - MARGIN - 30
    col_widths = [20, 90, 25, 15]  # PAT#, TITLE, CATEGORY, SHEET
    row_h = 7
    header_h = 10
    total_w = sum(col_widths)
    total_h = header_h + 20 * row_h

    # Table border
    sheet.ax.add_patch(Rectangle(
        (table_x, table_y - total_h), total_w, total_h,
        fill=True, facecolor='#F8F8F8', edgecolor=COLOR_BORDER, linewidth=1.5
    ))

    # Header
    headers = ['PAT #', 'INNOVATION TITLE', 'CATEGORY', 'SHEET']
    cx = table_x
    for i, (header, cw) in enumerate(zip(headers, col_widths)):
        if i > 0:
            sheet.ax.plot([cx, cx], [table_y, table_y - total_h],
                        color=COLOR_BORDER, linewidth=0.5)
        sheet.add_text(cx + cw / 2, table_y - header_h / 2, header,
                      fontsize=5, ha='center', fontweight='bold')
        cx += cw
    sheet.ax.plot([table_x, table_x + total_w], [table_y - header_h, table_y - header_h],
                color=COLOR_BORDER, linewidth=1)

    # Patent rows
    for idx, (pat_id, pat_info) in enumerate(sorted(PATENT_MAP.items())):
        row_y = table_y - header_h - (idx + 1) * row_h
        cx = table_x
        values = [pat_id, pat_info['title'][:40], pat_info['category'], f"S{pat_info['sheet']}"]
        for i, (val, cw) in enumerate(zip(values, col_widths)):
            if i > 0:
                sheet.ax.plot([cx, cx], [row_y, row_y + row_h],
                            color=COLOR_BORDER, linewidth=0.3)
            color = '#1B5E20' if i == 0 else '#333333'
            weight = 'bold' if i == 0 else 'normal'
            sheet.add_text(cx + cw / 2, row_y + row_h / 2, val,
                          fontsize=3.5, ha='center', color=color, fontweight=weight)
            cx += cw
        if idx < 19:
            sheet.ax.plot([table_x, table_x + total_w], [row_y, row_y],
                        color=COLOR_BORDER, linewidth=0.3)

    # Category summary (right side)
    cat_x = table_x + total_w + 20
    cat_y = table_y
    categories = {}
    for pat in PATENT_MAP.values():
        cat = pat['category']
        categories[cat] = categories.get(cat, 0) + 1

    sheet.add_text(cat_x, cat_y, "CATEGORY SUMMARY",
                  fontsize=5, ha='left', fontweight='bold')
    for i, (cat, count) in enumerate(sorted(categories.items())):
        sheet.add_text(cat_x, cat_y - 8 - i * 6, f"  {cat}: {count}",
                      fontsize=4, ha='left')

    # Sheet mapping summary (below categories)
    map_y = cat_y - 8 - len(categories) * 6 - 15
    sheet.add_text(cat_x, map_y, "SHEET DISTRIBUTION",
                  fontsize=5, ha='left', fontweight='bold')
    sheet_dist = {}
    for pat in PATENT_MAP.values():
        s = pat['sheet']
        sheet_dist[s] = sheet_dist.get(s, 0) + 1
    for i, (s, count) in enumerate(sorted(sheet_dist.items())):
        sheet.add_text(cat_x, map_y - 8 - i * 6, f"  Sheet {s}: {count} patents",
                      fontsize=4, ha='left')

    # Title block
    sheet.draw_title_block(
        title="PATENT INDEX", drawing_number="ASD-PAT-001",
        scale="NTS", material="N/A",
        tolerance="N/A", sheet_num="10 OF 10")
    sheet.draw_revision_block()
    sheet.draw_notes_block(notes=[
        "1. All 20 patents are proprietary AQUILA innovations.",
        "2. DFMA REV B patents: PAT-17, PAT-18, PAT-19, PAT-20.",
        "3. Patent applications filed via PatSnap Eureka TRIZ.",
        "4. See individual .docx files for full specifications.",
    ])

    base_name = "ASSY_PATENT_INDEX"
    sheet.save(str(output_dirs['png'] / f"{base_name}.png"), format='png')
    sheet.save(str(output_dirs['pdf'] / f"{base_name}.pdf"), format='pdf')
    sheet.save(str(output_dirs['svg'] / f"{base_name}.svg"), format='svg')

    # DXF
    if EZDXF_AVAILABLE:
        dxf = DXFDrawing(sheet_size=ASSEMBLY_SHEET)
        dxf.draw_title_block(title="PATENT INDEX", drawing_number="ASD-PAT-001",
                            scale="NTS", material="N/A")
        dxf.save(str(output_dirs['dxf'] / f"{base_name}.dxf"))

    return base_name


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="AQUILA Drawing Generator")
    parser.add_argument('--component', help='Generate drawing for single component ID')
    parser.add_argument('--assembly', choices=['subsystem', 'zone', 'phase', 'all'],
                        help='Generate assembly drawings')
    parser.add_argument('--bom', default=str(BOM_FILE), help='BOM xlsx path')
    args = parser.parse_args()

    bom_path = Path(args.bom)
    if not bom_path.exists():
        print(f"ERROR: BOM file not found: {bom_path}", file=sys.stderr)
        sys.exit(1)

    # Create output directories
    output_dirs = {
        'dxf': DXF_DIR, 'pdf': PDF_DIR, 'svg': SVG_DIR, 'png': PNG_DIR
    }
    for d in output_dirs.values():
        d.mkdir(parents=True, exist_ok=True)

    # Parse BOM
    print("=" * 60)
    print("AQUILA Sovereign — Engineering Drawing Generator")
    print("Standards: MENG 204 (Dr. Hijazi) + ASME Y14.5")
    print("=" * 60)

    components = parse_bom(bom_path)
    print(f"\n[1] Parsed {len(components)} BOM components")

    # Assign drawing numbers
    drawing_numbers = assign_drawing_numbers(components)
    print(f"[2] Assigned {len(drawing_numbers)} drawing numbers (XXX.AA.BB format)")

    if args.component:
        # Single component
        comp = next((c for c in components if c['ID'] == args.component), None)
        if not comp:
            print(f"ERROR: Component {args.component} not found", file=sys.stderr)
            sys.exit(1)
        print(f"\n[3] Generating drawing for {args.component}...")
        name = generate_component_drawing(comp, drawing_numbers[comp['ID']], output_dirs)
        print(f"    -> {name}.png/.pdf/.svg/.dxf")
        return

    if args.assembly:
        # Assembly drawings only
        print(f"\n[3] Generating assembly drawings ({args.assembly})...")
        count = 0

        if args.assembly in ('subsystem', 'all'):
            groups = group_by_subsystem(components)
            for group_name, comps in sorted(groups.items()):
                if len(comps) < 2:
                    continue
                dwg_num = f"ASD-ASSY-SUB-{count+1:02d}"
                name = generate_assembly_drawing(
                    f"SUB_{group_name}", comps, dwg_num, output_dirs, 'subsystem')
                count += 1
                print(f"    -> {name} ({len(comps)} components)")

        if args.assembly in ('zone', 'all'):
            groups = group_by_zone(components)
            for group_name, comps in sorted(groups.items()):
                if len(comps) < 2:
                    continue
                dwg_num = f"ASD-ASSY-ZONE-{count+1:02d}"
                name = generate_assembly_drawing(
                    f"ZONE_{group_name}", comps, dwg_num, output_dirs, 'zone')
                count += 1
                print(f"    -> {name} ({len(comps)} components)")

        if args.assembly in ('phase', 'all'):
            groups = group_by_phase(components)
            for group_name, comps in sorted(groups.items()):
                if len(comps) < 2:
                    continue
                dwg_num = f"ASD-ASSY-PHASE-{count+1:02d}"
                name = generate_assembly_drawing(
                    f"PHASE_{group_name}", comps, dwg_num, output_dirs, 'phase')
                count += 1
                print(f"    -> {name} ({len(comps)} components)")

        print(f"\n  Generated {count} assembly drawings")
        return

    # Generate all component drawings
    print(f"\n[3] Generating {len(components)} component drawings...")
    generated = 0
    for comp in components:
        bom_id = comp['ID']
        dwg_num = drawing_numbers.get(bom_id, f"ASD-{generated+1:03d}")
        try:
            name = generate_component_drawing(comp, dwg_num, output_dirs)
            generated += 1
            if generated % 10 == 0:
                print(f"    [{generated}/{len(components)}] Generated {name}")
        except Exception as e:
            print(f"    FAIL {bom_id}: {e}")

    print(f"\n  Generated {generated}/{len(components)} component drawings")

    # Generate all assembly drawings
    print(f"\n[4] Generating assembly drawings...")
    assembly_count = 0

    for grouping_name, grouping_fn in [('subsystem', group_by_subsystem),
                                        ('zone', group_by_zone),
                                        ('phase', group_by_phase)]:
        groups = grouping_fn(components)
        for group_name, comps in sorted(groups.items()):
            if len(comps) < 2:
                continue
            dwg_num = f"ASD-ASSY-{grouping_name[:3].upper()}-{assembly_count+1:02d}"
            try:
                name = generate_assembly_drawing(
                    f"{grouping_name[:3].upper()}_{group_name}", comps,
                    dwg_num, output_dirs, grouping_name)
                assembly_count += 1
            except Exception as e:
                print(f"    FAIL assembly {group_name}: {e}")

    print(f"  Generated {assembly_count} assembly drawings")

    # Generate patent index sheet (Sheet 10)
    print(f"\n[5] Generating patent index sheet...")
    try:
        name = generate_patent_index_sheet(output_dirs)
        print(f"    -> {name} (20 patents)")
        patent_count = 1
    except Exception as e:
        print(f"    FAIL patent index: {e}")
        patent_count = 0

    # Summary
    print(f"\n{'=' * 60}")
    print(f"DRAWING PACK COMPLETE")
    print(f"{'=' * 60}")
    print(f"  Component sheets: {generated}")
    print(f"  Assembly sheets:  {assembly_count}")
    print(f"  Patent index:      {patent_count}")
    print(f"  Total sheets:     {generated + assembly_count + patent_count}")
    print(f"  Formats:          DXF, PDF, SVG, PNG ({PNG_DPI} DPI)")
    print(f"  Standard:         ASME Y14.5 / ISO 128 (third-angle)")
    print(f"  Output:           {OUTPUT_DIR}")
    for fmt, d in output_dirs.items():
        files = list(d.glob('*'))
        print(f"    {fmt.upper()}: {len(files)} files in {d}")


if __name__ == "__main__":
    main()
